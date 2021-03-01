from django.shortcuts import render, redirect
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from student.models import *
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db.models import F
from .forms import *
from django.contrib import messages
from datetime import date


def student_register(request):
    if request.method == "POST":
        stufrm = StudentForm(request.POST)
        print(stufrm.errors)
        if stufrm.is_valid():
            roll1 = request.POST['Univroll']
            course1 = request.POST['c_name']
            batchs = request.POST['btch_start']
            batche = request.POST['btch_end']

            d = int(batche) - int(batchs)

            try:
                rl = Roll_no.objects.get(roll=roll1)
                print(course1)
                cs = course.objects.get(id=course1)
                aa = cs.total_year
                print(aa)
                print(d)
                print(date.today().year)
                if int(batchs) <= int(batche) and int(d) == int(aa) and int(batche) >= int(date.today().year):
                    stufrm.save()
                    messages.success(request, "Successfully Registered, please login")
                    return redirect('/Student_Login')

                else:
                    messages.error(request, "Enter the batches correctly according to course ")
                    c = course.objects.all()
                    return render(request, 'Register.html', context={'stufrm': stufrm, 'c': c})

            except Roll_no.DoesNotExist:
                messages.error(request, "University roll not matched from the database one")
                c = course.objects.all()
                return render(request, 'Register.html', context={'stufrm': stufrm, 'c': c})
        else:
            c = course.objects.all()
            return render(request, 'Register.html', context={'stufrm': stufrm, 'c': c})



    else:
        stufrm = StudentForm()
        c = course.objects.all()
        return render(request, 'Register.html', context={'stufrm': stufrm, 'c': c})


def stud_log(request):
    if not request.session.get('s_user_id', None):
        if request.method == "POST":
            SD = studentlogform(request.POST)
            Username = request.POST['s_username']
            Password = request.POST['s_password']
            print(SD.errors)
            try:
                rec = student.objects.get(Univroll=Username, alt_phone=Password)

                request.session['s_user_id'] = rec.id

                messages.success(request, "Successfully Logged in")
                return redirect('/Student_Dashboard')
            except student.DoesNotExist:

                c = course.objects.all()
                messages.error(request, "User not found or enter details correctly")

                return render(request, 'Student_Login.html', context={'SD': SD, 'c': c})

        else:
            SD = studentlogform()
            c = course.objects.all()
            return render(request, 'Student_Login.html', context={'SD': SD, 'c': c})
    else:
        messages.success(request, "already Logged in")
        return redirect('/Student_Dashboard')


def stud_out(request):
    if not request.session.get('s_user_id', None):
        messages.error(request, "please login first")
        return redirect('/Student_Login')
    else:
        del request.session['s_user_id']
        messages.success(request, "Successfully Logout")
        return redirect('/Student_Login')


def show_studentdata(request):
    if not request.session.get('s_user_id', None):
        messages.error(request, "Please Login first")
        return redirect('/Student_Login')
    else:
        i = student.objects.get(id=request.session['s_user_id'])

        return render(request, 'Student_Dashboard.html', {'i': i})


def show_not(request):
    if not request.session.get('s_user_id', None):
        messages.error(request, "Please Login first")
        return redirect('/Student_Login')
    else:
        i = student.objects.get(id=request.session['s_user_id'])
        l = assigned_addnot.objects.filter(
            Q(add_course=i.c_name, start_batch=i.btch_start, end_batch=i.btch_end) | Q(add_course=i.c_name,
                                                                                       start_batch=F(
                                                                                           'end_batch'))).order_by(
            '-d_te')
        return render(request, 'Student_Notification.html', {'l': l})


def stud_upd(request):
    if not request.session.get('s_user_id', None):
        messages.success(request, "please login first")
        return redirect('/Student_Login')
    else:
        i = student.objects.get(id=request.session['s_user_id'])
        fo = StudentForm(instance=i)
        if request.method == 'POST':
            fo = StudentForm(request.POST, instance=i)
            print(fo.errors)
            if fo.is_valid():
                fo.save()
                messages.success(request, "successfully updated")
                return redirect('/Student_Dashboard')

            else:
                messages.success(request, "enter details correctly")
                return render(request, 'Update_Details.html', context={'fo': fo, 'i': i})

        else:

            fo = StudentForm(instance=i)
            return render(request, 'Update_Details.html', context={'fo': fo, 'i': i})


def remove(request):
    stu = student.objects.get(id=request.session['s_user_id'])
    stu.delete()
    return redirect('assigned_dash.html')


# for placement memebrs

def base(request):
    c = course.objects.all()

    return render(request, 'Placement.html', {'c': c})


def plc_home(request, name):
    print(name)
    a = course.objects.get(cname=name)
    print(a)
    p = assigned.objects.get(COURSE_name=a)
    print(p)
    b = graph.objects.filter(C_name=a).order_by('-Batch_end')[:6]
    o = files.objects.filter(subjects=a)
    l = comp_graph.objects.filter(cx=a)
    m = plc_comp.objects.filter(companyfield=a)
    c = course.objects.all()

    return render(request, 'Course.html', context={'c': c, 'a': a, 'p': p, 'b': b, 'o': o, 'l': l, 'm': m})


def rolls(request):
    if request.method == "POST":
        cdfrm = rollnoform(request.POST)
        if cdfrm.is_valid():
            cou_rse = request.POST['c_course']
            s_batch = request.POST['batch_s']
            e_batch = request.POST['batch_e']
            froll = request.POST['f_roll']
            sroll = request.POST['s_roll']
            eroll = request.POST['l_roll']
            tot = request.POST['tots']

            rol = Workbook()

            rs = rol["Sheet"]
            course = str(cou_rse)
            Batch = str(s_batch) + str(e_batch)
            r1 = int(froll)
            r2 = int(sroll)
            last = int(eroll)
            print(type(last))
            dif = r2 - r1
            s = int(tot) + 1
            rs.cell(1, 1).value = "roll"
            i = 2
            while i <= s:
                rs.cell(i, 1).value = r1
                if r1 > last:
                    break
                else:
                    r1 = r1 + dif
                    i = i + 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment ; filename = data.xlsx'
            rol.save(response)
            return response
    else:
        cdfrm = rollnoform()
        return render(request, 'Download.html', {'cdfrm': cdfrm})
